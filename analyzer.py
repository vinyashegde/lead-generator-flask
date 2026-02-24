import os
import csv
import json
import uuid
import time
import requests
import re
from bs4 import BeautifulSoup
from google import genai
from google.genai import types
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_website_data(url):
    """Scrapes a website's text AND captures a full-page screenshot via Microlink."""
    if not url:
        return None
    
    data = {"text": "", "image_bytes": None}
    
    if not url.startswith('http'):
        url = 'http://' + url
        
    # Step 1: Capture a full-page scrolling screenshot using Microlink
    try:
        microlink_url = f"https://api.microlink.io/?url={url}&screenshot=true&meta=false&embed=screenshot.url&fullPage=true"
        img_response = requests.get(microlink_url, timeout=25)
        if img_response.status_code == 200 and len(img_response.content) > 5000:
            data["image_bytes"] = img_response.content
    except Exception:
        pass
        
    # Step 2: Scrape text content as fallback context
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/114.0.0.0 Safari/537.36'}
        response = requests.get(url, headers=headers, timeout=8)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            for script in soup(["script", "style"]):
                script.extract()
            text = soup.get_text(separator=' ', strip=True)
            data["text"] = text[:3000]
            return data
    except Exception:
        pass
    
    return data

def generate_website_pitch(scraped_data, api_key):
    """Uses Gemini 2.5 Flash Vision to generate an owner-facing agency pitch."""
    if not scraped_data:
        return ["Failed to extract website data."]
        
    try:
        client = genai.Client(api_key=api_key)
        
        prompt = f"""
        You are the Director of a Premium Web Development Agency. 
        You are writing a short, professional outreach pitch directly to the OWNER of this company's website.
        I am providing a full-page screenshot of their site and some text content.
        
        Your goal: Identify exactly 3 specific, technical issues with their website that hurt their business, and pitch how your agency would fix each one to boost their leads and sales.
        
        CRITICAL RULES:
        1. Each point MUST be a compelling, professional sentence aimed at a business owner (NOT a developer).
        2. MAXIMUM 20 words per point. Be concise and impactful.
        3. Focus on business impact: lost customers, slow loading, poor mobile experience, outdated design, missing trust signals, etc.
        4. Maintain a confident, consultative, premium agency tone.
        5. You MUST return your answer strictly as a valid JSON array of exactly 3 strings.
        
        Example:
        ["Your site takes over 5 seconds to load on mobile, causing 40% of potential customers to leave immediately.", "The homepage lacks a clear call-to-action, making it difficult for visitors to contact you or request a quote.", "Your website is not optimized for mobile devices, losing the 60% of users who browse on their phones."]
           
        DO NOT wrap the JSON in markdown code blocks. Just output the raw JSON array.
        
        Website Text: {scraped_data.get('text', '')[:2000]}
        """
        
        contents = [prompt]
        if scraped_data.get("image_bytes"):
            part = types.Part.from_bytes(data=scraped_data["image_bytes"], mime_type='image/png')
            contents.append(part)
        
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=contents,
        )
        
        raw_text = response.text.strip()
        
        # Strip potential markdown formatting
        if raw_text.startswith('```json'):
            raw_text = raw_text[7:]
        if raw_text.startswith('```'):
            raw_text = raw_text[3:]
        if raw_text.endswith('```'):
            raw_text = raw_text[:-3]
        raw_text = raw_text.strip()
        
        try:
            points_list = json.loads(raw_text)
            if isinstance(points_list, list) and len(points_list) > 0:
                # Clean each point
                cleaned = []
                for pt in points_list[:3]:
                    clean_pt = re.sub(r'[\r\n\*_~#]+', ' ', str(pt))
                    clean_pt = re.sub(r'\s+', ' ', clean_pt).strip()
                    cleaned.append(clean_pt)
                return cleaned
        except json.JSONDecodeError:
            pass
            
        # Fallback: return the raw text as a single-item list
        clean = re.sub(r'[\r\n\*_~#]+', ' ', raw_text)
        clean = re.sub(r'\s+', ' ', clean).strip()
        return [clean]
            
    except Exception as e:
        return [f"AI Generation Error: {str(e)}"]


def analyze_csv_file(input_filepath, gemini_api_key):
    """
    Generator that parses the CSV, scrapes URLs, calls Gemini, 
    and yields SSE events tracking progress. Outputs a styled .xlsx file.
    """
    if not os.path.exists(input_filepath):
        yield {"type": "error", "message": "Uploaded file not found."}
        return
        
    output_filepath = input_filepath.replace('.csv', '_analyzed.xlsx')
    
    yield {"type": "log", "message": "Starting Website Analysis Tool..."}
    
    try:
        with open(input_filepath, mode='r', encoding='utf-8') as infile:
            reader = csv.DictReader(infile)
            fieldnames = list(reader.fieldnames)
            if 'website' not in fieldnames:
                yield {"type": "error", "message": "CSV does not contain a 'website' column."}
                return
            
            # Add our analysis columns
            analysis_cols = ['Finding_1', 'Finding_2', 'Finding_3', 'Status']
            for col in analysis_cols:
                if col not in fieldnames:
                    fieldnames.append(col)
                
            rows = list(reader)
            total_rows = len(rows)
            yield {"type": "log", "message": f"Found {total_rows} leads to process."}
            
            # ── Process each row ──
            processed_rows = []
            processed_count = 0
            for row in rows:
                title = row.get('title', 'Unknown Business')
                website = row.get('website', '')
                
                if not website:
                    row['Finding_1'] = "No website found."
                    row['Finding_2'] = ""
                    row['Finding_3'] = ""
                    row['Status'] = "N/A"
                    yield {"type": "log", "message": f"Skipped {title}: No website."}
                else:
                    yield {"type": "log", "message": f"📸 Capturing screenshot for {title}..."}
                    scraped_data = extract_website_data(website)
                    
                    if scraped_data and (scraped_data.get("image_bytes") or scraped_data.get("text")):
                        yield {"type": "log", "message": f"🤖 Analyzing {title} with Gemini Vision..."}
                        points = generate_website_pitch(scraped_data, gemini_api_key)
                        row['Finding_1'] = points[0] if len(points) > 0 else ""
                        row['Finding_2'] = points[1] if len(points) > 1 else ""
                        row['Finding_3'] = points[2] if len(points) > 2 else ""
                        row['Status'] = "Pending Contact"
                        yield {"type": "log", "message": f"✅ Generated pitch for {title}."}
                    else:
                        row['Finding_1'] = "Could not access website."
                        row['Finding_2'] = ""
                        row['Finding_3'] = ""
                        row['Status'] = "Unreachable"
                        yield {"type": "log", "message": f"❌ Failed to scrape {website}."}
                        
                processed_rows.append(row)
                processed_count += 1
                
                yield {
                    "type": "progress",
                    "count": processed_count,
                    "total": total_rows,
                    "latest_lead": title
                }
                time.sleep(1)
            
            # ══════════════════════════════════════════════════
            # ── BUILD STYLED EXCEL WORKBOOK ──
            # ══════════════════════════════════════════════════
            yield {"type": "log", "message": "📊 Building styled Excel report..."}
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Website Analysis"
            
            # ── Style definitions ──
            header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            finding_fill_1 = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light gold
            finding_fill_2 = PatternFill(start_color='D6E8F7', end_color='D6E8F7', fill_type='solid')  # Light blue
            finding_fill_3 = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Light green
            
            status_pending_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # Light orange
            status_na_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')       # Light gray
            
            row_even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            cell_font = Font(name='Calibri', size=10)
            finding_font = Font(name='Calibri', size=10, color='333333')
            cell_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin', color='BFBFBF'),
                right=Side(style='thin', color='BFBFBF'),
                top=Side(style='thin', color='BFBFBF'),
                bottom=Side(style='thin', color='BFBFBF')
            )
            
            # ── Write header row ──
            for col_idx, header in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col_idx, value=header.replace('_', ' ').title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            # Freeze the header row
            ws.freeze_panes = 'A2'
            
            # ── Write data rows ──
            finding_cols = {fn: fieldnames.index(fn) + 1 for fn in ['Finding_1', 'Finding_2', 'Finding_3'] if fn in fieldnames}
            status_col = fieldnames.index('Status') + 1 if 'Status' in fieldnames else None
            
            for row_idx, row_data in enumerate(processed_rows, 2):
                is_even = row_idx % 2 == 0
                for col_idx, key in enumerate(fieldnames, 1):
                    value = row_data.get(key, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                    
                    # Alternate row shading for non-special columns
                    if is_even:
                        cell.fill = row_even_fill
                    
                    # Special coloring for finding columns
                    if col_idx == finding_cols.get('Finding_1') and value:
                        cell.fill = finding_fill_1
                        cell.font = finding_font
                    elif col_idx == finding_cols.get('Finding_2') and value:
                        cell.fill = finding_fill_2
                        cell.font = finding_font
                    elif col_idx == finding_cols.get('Finding_3') and value:
                        cell.fill = finding_fill_3
                        cell.font = finding_font
                    
                    # Special coloring for status column
                    if status_col and col_idx == status_col:
                        if value == 'Pending Contact':
                            cell.fill = status_pending_fill
                            cell.font = Font(name='Calibri', size=10, bold=True, color='C65911')
                        elif value in ('N/A', 'Unreachable'):
                            cell.fill = status_na_fill
                            cell.font = Font(name='Calibri', size=10, italic=True, color='808080')
            
            # ── Auto-size columns ──
            for col_idx, key in enumerate(fieldnames, 1):
                max_length = len(key) + 2
                for row_idx in range(2, len(processed_rows) + 2):
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                    max_length = max(max_length, min(len(cell_value), 50))
                
                # Finding columns are wider for readability
                if key.startswith('Finding'):
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 45
                elif key == 'Status':
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
                else:
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2
            
            # ── Set row heights for findings to breathe ──
            for row_idx in range(2, len(processed_rows) + 2):
                ws.row_dimensions[row_idx].height = 45
            ws.row_dimensions[1].height = 25
            
            # ── Save the workbook ──
            wb.save(output_filepath)
            
            filename_only = os.path.basename(output_filepath)
            yield {"type": "log", "message": f"✨ Analysis complete! Saved to {filename_only}"}
            yield {
                "type": "done",
                "filename": filename_only,
                "path": output_filepath,
                "count": total_rows
            }
            
    except Exception as e:
        yield {"type": "error", "message": f"Fatal Analyzer Error: {str(e)}"}
