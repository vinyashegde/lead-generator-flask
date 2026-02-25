import os
import re
import time
import datetime
import json
from serpapi import GoogleSearch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google import genai

def generate_smart_queries(goal, location, api_key, num_queries=3):
    """Uses Gemini to translate a business goal into precise Google X-Ray queries."""
    if not api_key:
        # Fallback to a basic search if no Gemini key is provided
        return [f'site:linkedin.com/in/ "{goal}" "{location}"']
        
    try:
        client = genai.Client(api_key=api_key)
        prompt = f"""
        Act as an Expert B2B Lead Generation Architect.
        The user wants to find LinkedIn leads based on this business goal:
        "{goal}"
        Location: "{location}"
        
        Generate EXACTLY {num_queries} highly specific Google X-Ray search strings targeting LinkedIn profiles (site:linkedin.com/in/).
        Target different angles:
        1. Decision makers (e.g., owner, founder, buyer)
        2. Specific niches mentioned (e.g., wellness clinic, holistic store)
        3. Relevant industry terms
        
        Format the output strictly as a JSON array of strings. Do not use markdown blocks.
        Example: ["site:linkedin.com/in/ \\"wellness center\\" \\"owner\\" \\"New York\\"", "site:linkedin.com/in/ \\"purchasing manager\\" \\"health tech\\"", "site:linkedin.com/in/ \\"biohacking\\" \\"founder\\""]
        """
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt,
        )
        raw = response.text.strip()
        if raw.startswith('```json'): raw = raw[7:]
        if raw.endswith('```'): raw = raw[:-3]
        
        queries = json.loads(raw.strip())
        return queries
    except Exception as e:
        print(f"Gemini query generation failed: {e}")
        return [f'site:linkedin.com/in/ "{goal}" "{location}"']

def evaluate_vendor_match(title, snippet, goal, api_key):
    """Uses Gemini to briefly explain why this lead matches the goal."""
    if not api_key or (not title and not snippet):
        return "Match evaluation unavailable."
        
    try:
        client = genai.Client(api_key=api_key)
        prompt = f"""
        Business Goal: "{goal}"
        Lead Profile: Title="{title}", Bio="{snippet}"
        
        Write a 1-sentence explanation of WHY this person is a good potential lead/vendor for the business goal.
        Be concise, direct, and persuasive. Max 15 words.
        """
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt,
        )
        return response.text.strip().replace('\n', ' ')
    except Exception:
        return "Potential profile match."

def convert_linkedin_leads_to_styled_xlsx(leads, output_xlsx_path):
    """Converts LinkedIn leads into a LinkedIn-branded styled Excel workbook."""
    if not leads:
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Smart Campaign Leads"
    
    fieldnames = list(leads[0].keys())
    
    # ── Style definitions (LinkedIn Blue Theme) ──
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0077B5', end_color='0077B5', fill_type='solid') # LinkedIn Blue
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Calibri', size=10)
    cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    row_even_fill = PatternFill(start_color='F3F6F8', end_color='F3F6F8', fill_type='solid') # Very light LinkedIn gray/blue
    
    # Highlight colors
    name_fill = PatternFill(start_color='D9EBF7', end_color='D9EBF7', fill_type='solid')
    match_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid') # Light green for match
    link_font = Font(name='Calibri', size=10, color='0077B5', underline='single')
    
    thin_border = Border(
        left=Side(style='thin', color='D9E1E7'),
        right=Side(style='thin', color='D9E1E7'),
        top=Side(style='thin', color='D9E1E7'),
        bottom=Side(style='thin', color='D9E1E7')
    )
    
    # ── Write header row ──
    display_names = {
        'name': 'Full Name',
        'title': 'Job Title',
        'location': 'Location',
        'link': 'LinkedIn Profile',
        'snippet': 'Profile Summary',
        'match_reason': 'Why they match?',
        'source_query': 'AI Search Query utilized'
    }
    
    for col_idx, key in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=display_names.get(key, key.replace('_', ' ').title()))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    ws.freeze_panes = 'A2'
    
    # ── Write data rows ──
    for row_idx, lead in enumerate(leads, 2):
        is_even = row_idx % 2 == 0
        for col_idx, key in enumerate(fieldnames, 1):
            value = lead.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            
            if is_even:
                cell.fill = row_even_fill
            
            if key == 'name':
                cell.fill = name_fill
                cell.font = Font(name='Calibri', size=10, bold=True)
            elif key == 'link':
                cell.font = link_font
            elif key == 'match_reason':
                cell.fill = match_fill
                cell.font = Font(name='Calibri', size=10, italic=True)
    
    # ── Auto-size columns ──
    col_widths = {
        'name': 25, 'title': 30, 'location': 20, 'link': 45, 'match_reason': 50, 'snippet': 40, 'source_query': 35
    }
    for col_idx, key in enumerate(fieldnames, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(key, 20)
    
    # Row heights
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, len(leads) + 2):
        ws.row_dimensions[row_idx].height = 40 # Increased height for match reason wrapper
    
    wb.save(output_xlsx_path)

def generate_linkedin_leads(goal, location, limit, api_key, gemini_api_key=None):
    """
    Fetch LinkedIn leads using AI-generated Smart Campaign queries via SerpAPI.
    Yields events for SSE streaming.
    """
    leads = []
    seen_links = set()
    
    os.makedirs('generated_leads', exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_goal = re.sub(r'[^a-zA-Z0-9]', '_', goal)[:20]
    safe_location = re.sub(r'[^a-zA-Z0-9]', '_', location)[:15]
    xlsx_filename = f"smart_campaign_{safe_goal}_{safe_location}_{timestamp}.xlsx"
    xlsx_output_path = os.path.join("generated_leads", xlsx_filename)
    
    yield {"type": "log", "message": f"🤖 Analyzing Business Goal: '{goal}'..."}
    
    queries = generate_smart_queries(goal, location, gemini_api_key, num_queries=3)
    yield {"type": "log", "message": f"🧠 Generated {len(queries)} specific X-Ray permutations."}
    
    for query in queries:
        if len(leads) >= limit:
            break
            
        yield {"type": "log", "message": f"🔍 Running query: {query}"}
        
        start = 0
        while len(leads) < limit and start < 30: # Max 3 pages per query to keep it diverse
            params = {
                "engine": "google",
                "q": query,
                "api_key": api_key,
                "start": start,
                "num": 10
            }
            
            try:
                search = GoogleSearch(params)
                results = search.get_dict()
                organic_results = results.get("organic_results", [])
                
                if not organic_results:
                    break # Move to next query
                
                for result in organic_results:
                    if len(leads) >= limit:
                        break
                        
                    link = result.get("link", "")
                    if "/in/" not in link or link in seen_links:
                        continue
                    
                    seen_links.add(link)
                    
                    raw_title = result.get("title", "Unknown")
                    name = raw_title.split('-')[0].split('|')[0].strip()
                    
                    job_title = "N/A"
                    if '-' in raw_title:
                        parts = raw_title.split('-')
                        if len(parts) > 1:
                            job_title = parts[1].split('|')[0].strip()
                    
                    snippet = result.get("snippet", "")
                    
                    # AI Match Analysis
                    yield {"type": "log", "message": f"⚙️ Evaluating match: {name}"}
                    match_reason = evaluate_vendor_match(job_title, snippet, goal, gemini_api_key)
                    
                    lead = {
                        "name": name,
                        "title": job_title,
                        "location": location,
                        "link": link,
                        "match_reason": match_reason,
                        "snippet": snippet,
                        "source_query": query
                    }
                    
                    leads.append(lead)
                    
                    yield {
                        "type": "progress",
                        "count": len(leads),
                        "total": limit,
                        "latest_lead": name
                    }
                
                start += 10
                time.sleep(1)
                
            except Exception as e:
                yield {"type": "error", "message": f"LinkedIn Search Error: {str(e)}"}
                return
                
    if leads:
        yield {"type": "log", "message": "📊 Formatting Smart Campaign Report..."}
        try:
            convert_linkedin_leads_to_styled_xlsx(leads, xlsx_output_path)
        except Exception as e:
            yield {"type": "error", "message": f"Excel generation failed: {str(e)}"}
            return
            
    yield {"type": "log", "message": f"✨ Success! Compiled {len(leads)} targeted LinkedIn leads."}
    yield {
        "type": "done",
        "filename": xlsx_filename,
        "path": xlsx_output_path,
        "count": len(leads)
    }
