import os
import csv
import time
import re
import requests
import json
from bs4 import BeautifulSoup
from serpapi import GoogleSearch
from urllib.parse import urlparse
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_emails_from_text(text):
    """Extract emails from text using regex"""
    email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    emails = set(re.findall(email_pattern, text))
    # Filter out invalid emails that might be image filenames or too long
    valid_emails = {e for e in emails if len(e) < 50 and not any(ext in e.lower() for ext in ['.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp'])}
    return valid_emails

def scrape_email_from_website(url, emit_log=None):
    """Scrape emails from a website"""
    if not url:
        return ""
    
    try:
        # Add http if missing
        if not url.startswith('http'):
            url = 'http://' + url
            
        if emit_log:
            emit_log(f"Scraping emails from {url}...")
            
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            # Use separator to avoid concatenated text like "info@kykindia.comHomeAbout"
            text = soup.get_text(separator=' ')
            emails = extract_emails_from_text(text)
            
            # Also check contact page if found
            if not emails:
                contact_links = soup.find_all('a', href=re.compile(r'contact', re.I))
                for link in contact_links:
                    contact_url = link.get('href')
                    if contact_url:
                        if not contact_url.startswith('http'):
                            base_url = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
                            contact_url = base_url + contact_url if contact_url.startswith('/') else base_url + '/' + contact_url
                        
                        try:
                            contact_response = requests.get(contact_url, headers=headers, timeout=5)
                            if contact_response.status_code == 200:
                                contact_soup = BeautifulSoup(contact_response.text, 'html.parser')
                                emails.update(extract_emails_from_text(contact_soup.get_text(separator=' ')))
                        except:
                            pass
            
            # check mailto links
            mailto_links = soup.select('a[href^=mailto]')
            for link in mailto_links:
                href = link.get('href', '')
                email = href.replace('mailto:', '').split('?')[0].strip()
                if email:
                    emails.add(email)

            filtered_emails = {e for e in emails if not any(x in e.lower() for x in ['example.com', 'yourdomain', 'sentry.io', 'wixpress.com'])}
            
            if filtered_emails:
                result = ", ".join(filtered_emails)
                if emit_log:
                    emit_log(f"Found emails: {result}")
                return result
    except Exception as e:
        if emit_log:
            emit_log(f"Scrape error for {url}: {str(e)[:50]}")
        pass
        
    return ""


def convert_leads_to_styled_xlsx(leads, output_xlsx_path):
    """Converts a list of lead dicts into a beautifully styled Excel workbook."""
    if not leads:
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Generated Leads"
    
    fieldnames = list(leads[0].keys())
    
    # ── Style definitions ──
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Calibri', size=10)
    cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    row_even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Highlight colors for key columns
    email_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')      # Light green
    website_fill = PatternFill(start_color='D6E8F7', end_color='D6E8F7', fill_type='solid')    # Light blue
    phone_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')      # Light gold
    rating_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')     # Light orange
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # ── Write header row ──
    display_names = {
        'title': 'Business Name',
        'address': 'Address',
        'phone': 'Phone',
        'website': 'Website',
        'email': 'Email',
        'rating': 'Rating',
        'reviews': 'Reviews',
        'type': 'Category',
        'source_query': 'Search Query'
    }
    
    for col_idx, key in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=display_names.get(key, key.replace('_', ' ').title()))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    ws.freeze_panes = 'A2'
    
    # ── Column index mapping for special coloring ──
    col_map = {key: idx + 1 for idx, key in enumerate(fieldnames)}
    
    # ── Write data rows ──
    for row_idx, lead in enumerate(leads, 2):
        is_even = row_idx % 2 == 0
        for col_idx, key in enumerate(fieldnames, 1):
            value = lead.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            
            if is_even:
                cell.fill = row_even_fill
            
            # Special column coloring when cell has data
            if key == 'email' and value:
                cell.fill = email_fill
                cell.font = Font(name='Calibri', size=10, color='375623')
            elif key == 'website' and value:
                cell.fill = website_fill
                cell.font = Font(name='Calibri', size=10, color='1F4E79')
            elif key == 'phone' and value:
                cell.fill = phone_fill
            elif key == 'rating' and value:
                cell.fill = rating_fill
                cell.font = Font(name='Calibri', size=10, bold=True, color='C65911')
    
    # ── Auto-size columns ──
    col_widths = {
        'title': 30, 'address': 35, 'phone': 16, 'website': 30,
        'email': 30, 'rating': 10, 'reviews': 10, 'type': 20, 'source_query': 25
    }
    for col_idx, key in enumerate(fieldnames, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(key, 18)
    
    # Row heights
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, len(leads) + 2):
        ws.row_dimensions[row_idx].height = 28
    
    wb.save(output_xlsx_path)


def generate_leads(keyword, location, limit, api_key, require_email=False, require_website=False):
    """
    Fetch leads using SerpAPI Google Maps search.
    Yields progress dictionaries for Server-Sent Events (SSE).
    """
    leads = []
    seen_identifiers = set()
    
    # Setup directory and filename
    os.makedirs('generated_leads', exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = re.sub(r'[^a-zA-Z0-9]', '_', keyword)
    safe_location = re.sub(r'[^a-zA-Z0-9]', '_', location)
    xlsx_filename = f"leads_{safe_keyword}_{safe_location}_{timestamp}.xlsx"
    xlsx_output_path = os.path.join("generated_leads", xlsx_filename)
    
    yield {"type": "log", "message": f"Starting lead generation for '{keyword}' in '{location}'. Target: {limit} leads."}
    
    query = f"{keyword} {location}"
    start = 0
    
    while len(leads) < limit:
        yield {"type": "log", "message": f"Fetching results from SerpAPI (start={start})..."}
        params = {
            "engine": "google_maps",
            "q": query,
            "type": "search",
            "api_key": api_key,
            "start": start
        }
        
        try:
            search = GoogleSearch(params)
            results = search.get_dict()
            local_results = results.get("local_results", [])
            
            if not local_results:
                yield {"type": "log", "message": "No more results found from Google Maps."}
                break
            
            for result in local_results:
                if len(leads) >= limit:
                    break
                    
                title = result.get("title")
                phone = result.get("phone")
                address = result.get("address")
                website = result.get("website")
                
                dedup_key = phone if phone else (f"{title}|{address}" if address else title)
                
                if dedup_key and dedup_key not in seen_identifiers:
                    seen_identifiers.add(dedup_key)
                    
                    email = ""
                    if website:
                        yield {"type": "log", "message": f"Scraping emails from {website}..."}
                        email = scrape_email_from_website(website)
                        if email:
                            yield {"type": "log", "message": f"Found emails: {email}"}
                    
                    if require_website and not website:
                        yield {"type": "log", "message": f"Skipping {title} (no website)"}
                        continue
                        
                    if require_email and not email:
                        yield {"type": "log", "message": f"Skipping {title} (no email)"}
                        continue
                    
                    lead = {
                        "title": title,
                        "address": address,
                        "phone": phone,
                        "website": website,
                        "email": email,
                        "rating": result.get("rating"),
                        "reviews": result.get("reviews"),
                        "type": result.get("type"),
                        "source_query": query,
                    }
                    
                    leads.append(lead)
                    
                    # Emit progress update
                    yield {
                        "type": "progress",
                        "count": len(leads),
                        "total": limit,
                        "latest_lead": title
                    }
            
            start += 20 # Google maps pagination typically goes by 20
            time.sleep(1) # Be nice to SerpAPI
            
        except Exception as e:
            yield {"type": "error", "message": f"Google Maps/SerpAPI Error: {str(e)}"}
            return
    
    # ── Build styled Excel workbook ──
    if leads:
        yield {"type": "log", "message": "📊 Building styled Excel report..."}
        try:
            convert_leads_to_styled_xlsx(leads, xlsx_output_path)
        except Exception as e:
            yield {"type": "error", "message": f"Excel generation error: {str(e)}"}
            return
    
    yield {"type": "log", "message": f"✨ Finished! Generated {len(leads)} leads. Saved to {xlsx_filename}"}
    
    yield {
        "type": "done",
        "filename": xlsx_filename,
        "path": xlsx_output_path,
        "count": len(leads)
    }

