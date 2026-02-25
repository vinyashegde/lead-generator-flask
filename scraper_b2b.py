import os
import re
import time
import datetime
from bs4 import BeautifulSoup
import requests
from serpapi import GoogleSearch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_email_and_phone_from_url(url):
    """Scrapes a given URL to find any email addresses or phone-like numbers."""
    if not url:
        return "N/A", "N/A"
        
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/114.0.0.0 Safari/537.36'}
        response = requests.get(url, headers=headers, timeout=5)
        
        if response.status_code == 200:
            text = response.text
            soup = BeautifulSoup(text, 'html.parser')
            clean_text = soup.get_text()
            
            # Look for mailto links first (most reliable)
            email = "N/A"
            for a in soup.find_all('a', href=True):
                if a['href'].startswith('mailto:'):
                    email = a['href'].replace('mailto:', '').split('?')[0].strip()
                    break
            
            # Fallback to regex if no mailto found
            if email == "N/A":
                email_match = re.search(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+', clean_text)
                if email_match:
                    email = email_match.group(0)
                    
            # Basic phone number regex (US focus but catches general formats)
            phone = "N/A"
            phone_match = re.search(r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', clean_text)
            if phone_match:
                phone = phone_match.group(0)
                
            return email, phone
            
    except requests.exceptions.RequestException:
        pass
        
    return "N/A", "N/A"

def find_decision_maker_on_linkedin(company_name, api_key):
    """Uses Google X-Ray to find the Owner or Buyer of a specific company on LinkedIn."""
    if not api_key:
        return "N/A", "N/A", "N/A"
        
    query = f'site:linkedin.com/in/ "{company_name}" ("owner" OR "founder" OR "buyer" OR "purchasing")'
    params = {
        "engine": "google",
        "q": query,
        "api_key": api_key,
        "num": 1
    }
    
    try:
        search = GoogleSearch(params)
        results = search.get_dict()
        organic_results = results.get("organic_results", [])
        
        if organic_results:
            top_result = organic_results[0]
            link = top_result.get("link", "")
            
            if "/in/" in link:
                raw_title = top_result.get("title", "Unknown")
                name = raw_title.split('-')[0].split('|')[0].strip()
                
                # Try to extract their exact job title from the snippet
                snippet = top_result.get("snippet", "")
                
                return name, link, snippet
    except Exception:
        pass
        
    return "Not Found", "N/A", "N/A"

def convert_b2b_leads_to_styled_xlsx(leads, output_xlsx_path):
    """Converts B2B leads into a premium green/gold styled Excel workbook."""
    if not leads:
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "B2B Partners"
    
    fieldnames = list(leads[0].keys())
    
    # ── Style definitions (Premium Green Theme for Business) ──
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E3A2F', end_color='1E3A2F', fill_type='solid') # Dark Green
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Calibri', size=10)
    cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    row_even_fill = PatternFill(start_color='F4F7F6', end_color='F4F7F6', fill_type='solid')
    
    # Highlight colors
    company_fill = PatternFill(start_color='E8F0EA', end_color='E8F0EA', fill_type='solid')
    human_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid') # Soft Gold for human
    link_font = Font(name='Calibri', size=10, color='0077B5', underline='single')
    
    thin_border = Border(
        left=Side(style='thin', color='D9E1E7'), right=Side(style='thin', color='D9E1E7'),
        top=Side(style='thin', color='D9E1E7'), bottom=Side(style='thin', color='D9E1E7')
    )
    
    # ── Write header row ──
    display_names = {
        'company_name': 'Business Name',
        'company_website': 'Website',
        'company_email': 'Official Email',
        'company_phone': 'Official Phone',
        'rating': 'Google Rating',
        'address': 'Address',
        'decision_maker_name': 'Owner / Buyer Name',
        'decision_maker_linkedin': 'Owner LinkedIn',
        'decision_maker_bio': 'Owner Bio / Context'
    }
    
    for col_idx, key in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=display_names.get(key, key.replace('_', ' ').title()))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    ws.freeze_panes = 'A2'
    
    # ── Write data rows ──
    for row_idx, lead in enumerate(leads, 2):
        is_even = row_idx % 2 == 0
        for col_idx, key in enumerate(fieldnames, 1):
            value = lead.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            
            if is_even:
                cell.fill = row_even_fill
            
            if key == 'company_name':
                cell.fill = company_fill
                cell.font = Font(name='Calibri', size=10, bold=True)
            elif key == 'decision_maker_name':
                cell.fill = human_fill
                cell.font = Font(name='Calibri', size=10, bold=True)
            elif 'website' in key or 'linkedin' in key:
                if value and value != 'N/A' and value != 'Not Found':
                    cell.font = link_font
    
    # ── Auto-size columns ──
    col_widths = {
        'company_name': 25, 'company_website': 25, 'company_email': 25, 'company_phone': 15,
        'rating': 12, 'address': 30, 'decision_maker_name': 20, 'decision_maker_linkedin': 35, 'decision_maker_bio': 40
    }
    for col_idx, key in enumerate(fieldnames, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(key, 20)
    
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, len(leads) + 2):
        ws.row_dimensions[row_idx].height = 30
    
    wb.save(output_xlsx_path)

def generate_b2b_partners(keyword, location, limit, api_key):
    """
    Hybrid Scraper: Finds local businesses, scrapes their websites for emails, 
    then uses LinkedIn X-Ray to find the exact owner/buyer.
    Yields events for SSE streaming.
    """
    leads = []
    
    os.makedirs('generated_leads', exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = re.sub(r'[^a-zA-Z0-9]', '_', keyword)
    safe_location = re.sub(r'[^a-zA-Z0-9]', '_', location)
    xlsx_filename = f"b2b_partners_{safe_keyword}_{safe_location}_{timestamp}.xlsx"
    xlsx_output_path = os.path.join("generated_leads", xlsx_filename)
    
    yield {"type": "log", "message": f"🏢 Stage 1: Finding physical '{keyword}' businesses in '{location}'..."}
    
    search_query = f"{keyword} in {location}"
    params = {
        "engine": "google_maps",
        "q": search_query,
        "type": "search",
        "api_key": api_key,
    }
    
    try:
        search = GoogleSearch(params)
        results = search.get_dict()
        local_results = results.get("local_results", [])
        
        if not local_results:
            yield {"type": "error", "message": "No businesses found on Google Maps. Try broader terms like 'Wellness Center', 'Health Food Store', or 'Water Purification'."}
            return
            
        for result in local_results:
            if len(leads) >= limit:
                break
                
            company_name = result.get("title", "")
            if not company_name:
                continue
                
            website = result.get("website", "")
            phone = result.get("phone", "N/A")
            rating = str(result.get("rating", "N/A"))
            address = result.get("address", "N/A")
            
            yield {"type": "log", "message": f"🌐 Found Business: {company_name}"}
            
            # Scrape Website for Email
            email = "N/A"
            if website:
                yield {"type": "log", "message": f"   Scraping website for email: {website}"}
                extracted_email, extracted_phone = extract_email_and_phone_from_url(website)
                email = extracted_email
                if phone == "N/A":
                    phone = extracted_phone
            else:
                website = "N/A"
                
            # Cross-reference LinkedIn for the human decision maker
            yield {"type": "log", "message": f"   🕵️‍♂️ X-Ray searching LinkedIn for the Owner/Buyer..."}
            dm_name, dm_link, dm_bio = find_decision_maker_on_linkedin(company_name, api_key)
            
            lead = {
                "company_name": company_name,
                "company_website": website,
                "company_email": email,
                "company_phone": phone,
                "rating": rating,
                "address": address,
                "decision_maker_name": dm_name,
                "decision_maker_linkedin": dm_link,
                "decision_maker_bio": dm_bio
            }
            
            leads.append(lead)
            
            yield {
                "type": "progress",
                "count": len(leads),
                "total": limit,
                "latest_lead": company_name
            }
            
            time.sleep(1) # Be respectful
            
    except Exception as e:
        yield {"type": "error", "message": f"B2B Search Error: {str(e)}"}
        return
        
    if leads:
        yield {"type": "log", "message": "📊 Formatting B2B Partnership Dossier..."}
        try:
            convert_b2b_leads_to_styled_xlsx(leads, xlsx_output_path)
        except Exception as e:
            yield {"type": "error", "message": f"Excel generation failed: {str(e)}"}
            return
            
    yield {"type": "log", "message": f"✨ Success! Compiled {len(leads)} B2B Vendor Partnerships."}
    yield {
        "type": "done",
        "filename": xlsx_filename,
        "path": xlsx_output_path,
        "count": len(leads)
    }
